# conto_economico_app.py
from flask import Flask, render_template_string, request, send_file, redirect, url_for
import io, re, math
import pandas as pd
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# -------------------- util --------------------
def parse_num(v, default=0.0):
    if v is None:
        return default
    v = str(v).strip().replace(" ", "").replace(",", ".")
    try:
        return float(v)
    except:
        return default

def slugify(text, default="operazione"):
    if not text:
        return default
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or default

def amortize_first_year(principal, annual_rate_pct, years):
    """Restituisce (rata_mensile, esborso_annuo, interesse_anno1, capitale_anno1, residuo_fine_anno1)."""
    P = float(principal)
    i = float(annual_rate_pct) / 100.0
    n_years = int(years) if years else 0
    if P <= 0 or i < 0 or n_years <= 0:
        return (0.0, 0.0, 0.0, 0.0, P)

    r = i / 12.0
    N = n_years * 12
    if r == 0:
        rata = P / N
    else:
        rata = P * (r) / (1 - (1 + r) ** (-N))

    saldo = P
    interesse_anno = 0.0
    capitale_anno = 0.0
    for _ in range(12):
        interesse = saldo * r
        quota_cap = rata - interesse
        interesse_anno += interesse
        capitale_anno += quota_cap
        saldo -= quota_cap
        if saldo < 0:
            saldo = 0.0
            break

    esborso_annuo = rata * 12
    return (rata, esborso_annuo, interesse_anno, capitale_anno, saldo)

# -------------------- template --------------------
HTML = """
<!doctype html>
<html lang="it">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Calcolo Conto Economico</title>
<style>
  :root{ --bg:#0f172a; --card:#111827; --muted:#94a3b8; --accent:#22c55e; --txt:#e5e7eb; --ring:#374151; }
  body{ margin:0; background:linear-gradient(180deg,#0b1023,#0e1227 40%,#0f172a); color:var(--txt);
        font-family:system-ui,-apple-system,Segoe UI,Roboto,Inter,Helvetica,Arial }
  header{text-align:center;padding:16px}
  h1{margin:0;font-size:24px}
  .wrap{max-width:1100px;margin:0 auto;padding:20px}
  .card{background:rgba(17,24,39,.85);border:1px solid #1f2937;border-radius:14px;
        padding:16px;box-shadow:0 6px 20px rgba(0,0,0,.25);margin-bottom:16px}
  .card.results{background:#1e293b;}

  /* Parent segmented control */
  .parent-tabs{display:flex;justify-content:center;margin:14px auto 0;flex-wrap:wrap}
  .segment{
    display:flex;gap:0;background:#0b1222;border:1px solid #1f2937;border-radius:12px;
    padding:4px;box-shadow:inset 0 0 0 1px #0b1222
  }
  .parent-btn{
    background:transparent;border:none;color:#94a3b8;font-weight:700;font-size:14px;
    padding:8px 14px;border-radius:8px;cursor:pointer;transition:all .2s
  }
  .parent-btn:hover{color:#e5e7eb}
  .parent-btn.active{
    background:#17243a;color:#e5e7eb;box-shadow:0 0 0 1px #243b55 inset
  }

  /* Child tabs */
  .tabs{display:flex;flex-wrap:wrap;justify-content:center;gap:30px;margin:12px 0 16px;border-bottom:2px solid #1f2937}
  .tablink{background:none;border:none;color:#94a3b8;font-weight:600;font-size:15px;
           padding:10px 0;cursor:pointer;position:relative;transition:color .25s}
  .tablink:hover{color:#e5e7eb}
  .tablink.active{color:var(--accent)}
  .tablink.active::after{content:"";position:absolute;bottom:-2px;left:0;right:0;height:3px;background:var(--accent);border-radius:2px}
  .tabcontent{display:none}

  .row{margin:8px 0;display:flex;gap:10px;align-items:center}
  .row label{flex:1}
  .row input, .row select{flex:1;padding:10px;border-radius:10px;border:1px solid var(--ring);background:#0b1023;color:var(--txt)}

  .actions{display:flex;gap:10px;justify-content:flex-end;margin-top:12px;flex-wrap:wrap}
  .btn{padding:8px 14px;border:none;border-radius:10px;cursor:pointer;font-weight:700;text-decoration:none;display:inline-block;text-align:center}
  .primary{background:linear-gradient(180deg,#16a34a,#15803d);color:#fff}
  .secondary{background:#0b1222;color:#e5e7eb;border:1px solid #1f2937}

  .pill{background:#0b1222;border:1px solid #1f2937;border-radius:12px;padding:10px;margin:6px 0}
  .pill.roi-good { background:#14532d; border:1px solid #15803d; color:#d1fae5; }

  .grid3{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:10px}

  .preview{margin-top:10px;padding:12px;border-radius:12px;background:#0b1222;border:1px solid #1f2937}
  .preview h4{margin:0 0 8px;font-size:14px;color:#cbd5e1}
  .preview .item{display:flex;justify-content:space-between;margin:4px 0;font-size:14px}
  .preview .item span:first-child{color:#a7b2c3}

  .muted{color:#94a3b8;font-size:13px}

  table{width:100%;border-collapse:collapse;margin-top:8px}
  th,td{border:1px solid #1f2937;padding:8px;text-align:right}
  th{text-align:center;background:#0b1222}
  td:first-child, th:first-child{text-align:left}
</style>
</head>
<body>
<header>
  <h1>Calcolo Conto Economico</h1>
  <p>Strumento per analisi investimenti immobiliari</p>
</header>

<div class="wrap card">
  <!-- Parent tabs -->
  <div class="parent-tabs">
    <div class="segment">
      <button type="button" class="parent-btn" data-parent="compravendita" onclick="openParent('compravendita')">Compravendita</button>
      <button type="button" class="parent-btn" data-parent="affitti" onclick="openParent('affitti')">Affitti</button>
      <button type="button" class="parent-btn" data-parent="mutuo" onclick="openParent('mutuo')">Mutuo</button>
    </div>
  </div>
</div>

<!-- ========================== COMPRAVENDITA ========================== -->
<div class="wrap card parent-section" id="parent-compravendita">
  <form method="post" id="form_compravendita">
    <input type="hidden" name="parent_tab" value="compravendita"/>
    <input type="hidden" id="active_tab_cv" name="active_tab" value="{{ active_tab_cv or 'cv_acq' }}"/>

    <div class="tabs">
      <button type="button" class="tablink" data-tab="cv_acq" onclick="openChild('compravendita','cv_acq')">Acquisto</button>
      <button type="button" class="tablink" data-tab="cv_cat" onclick="openChild('compravendita','cv_cat')">Valore Catastale & Registro</button>
      <button type="button" class="tablink" data-tab="cv_vend" onclick="openChild('compravendita','cv_vend')">Costi messa in vendita</button>
      <button type="button" class="tablink" data-tab="cv_val" onclick="openChild('compravendita','cv_val')">Nuovo Valore</button>
    </div>

    <!-- CV: ACQUISTO -->
    <div id="cv_acq" class="tabcontent">
      <div class="row"><label>Titolo operazione</label><input type="text" name="cv_titolo" value="{{ cv.titolo }}"></div>

      <div class="row"><label>Acquisto immobile</label><input type="text" name="cv_ask" value="{{ cv.ask }}"></div>
      <div class="row"><label>Imposta ipotecaria</label><input type="text" name="cv_ipotecaria" value="{{ cv.ipotecaria }}"></div>
      <div class="row"><label>Imposta catastale</label><input type="text" name="cv_catastale" value="{{ cv.catastale }}"></div>
      <div class="row"><label>Imposta di registro</label><input type="text" value="{{ cv.imposta_registro }}" readonly></div>
      <div class="row"><label>Provvigioni agenzia (acquisto)</label><input type="text" name="cv_agenzia" value="{{ cv.agenzia }}"></div>
      <div class="row"><label>Studio architetto</label><input type="text" name="cv_architetto" value="{{ cv.architetto }}"></div>
      <div class="row"><label>Condono</label><input type="text" name="cv_condono" value="{{ cv.condono }}"></div>
      <div class="row"><label>Spese condominiali insolute</label><input type="text" name="cv_condominio" value="{{ cv.condominio }}"></div>
      <div class="row"><label>Nuove utenze (luce+gas)</label><input type="text" name="cv_utenze" value="{{ cv.utenze }}"></div>
      <div class="row"><label>Imprevisti</label><input type="text" name="cv_imprevisti" value="{{ cv.imprevisti }}"></div>

      <div class="row">
        <label>Tipo di ristrutturazione</label>
        <select name="cv_ristrut_tipo">
          <option value="nessuna"   {% if cv.ristrut_tipo=='nessuna' %}selected{% endif %}>Nessuna</option>
          <option value="piccola"   {% if cv.ristrut_tipo=='piccola' %}selected{% endif %}>Piccoli interventi (10%)</option>
          <option value="intermedia"{% if cv.ristrut_tipo=='intermedia' %}selected{% endif %}>Ristrutturazione intermedia (20%)</option>
          <option value="complessa" {% if cv.ristrut_tipo=='complessa' %}selected{% endif %}>Ristrutturazione complessa (60%)</option>
        </select>
      </div>
      <div class="row"><label>Costo ristrutturazione (calcolato)</label><input type="text" value="{{ cv.ristrutturazione }}" readonly></div>

      <div class="actions">
        <button class="btn primary" type="submit">Calcola</button>
        <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
      </div>
    </div>

    <!-- CV: CATASTALE -->
    <div id="cv_cat" class="tabcontent">
      <div class="row">
        <label>Tipo di proprietà</label>
        <select name="cv_tipo" id="cv_tipo" onchange="autoFillByTipo('cv');updatePreviewCat('cv');">
          <option value="prima"  {{ 'selected' if cv.tipo=='prima' else '' }}>Prima casa</option>
          <option value="seconda"{{ 'selected' if cv.tipo=='seconda' else '' }}>Seconda casa</option>
        </select>
      </div>
      <div class="row"><label>Rendita catastale</label><input type="text" id="cv_rendita" name="cv_rendita" value="{{ cv.rendita }}" oninput="updatePreviewCat('cv')"></div>
      <div class="row"><label>Coefficiente</label><input type="text" id="cv_coeff" name="cv_coeff" value="{{ cv.coeff }}" oninput="updatePreviewCat('cv')"></div>
      <div class="row"><label>Imposta di registro %</label><input type="text" id="cv_imposta_pct" name="cv_imposta_pct" value="{{ cv.imposta_pct }}" oninput="updatePreviewCat('cv')"></div>

      <div class="preview">
        <h4>Anteprima calcoli</h4>
        <div class="item"><span>Valore catastale</span><span id="cv_pv_val_cat">—</span></div>
        <div class="item"><span>Imposta di registro</span><span id="cv_pv_imp_reg">—</span></div>
      </div>
      <div class="actions">
        <button class="btn primary" type="submit">Calcola</button>
        <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
      </div>
    </div>

    <!-- CV: VENDITA -->
    <div id="cv_vend" class="tabcontent">
      <div class="row">
        <label>Home staging % (sul prezzo di vendita)</label>
        <select name="cv_hs_percent" onchange="updatePreviewVendita('cv')">
          <option value="1" {% if cv.hs_percent=='1' %}selected{% endif %}>1%</option>
          <option value="2" {% if cv.hs_percent=='2' %}selected{% endif %}>2%</option>
          <option value="3" {% if cv.hs_percent=='3' %}selected{% endif %}>3%</option>
        </select>
      </div>
      <div class="row"><label>APE</label><input type="text" id="cv_ape" name="cv_ape" value="{{ cv.ape }}" oninput="updatePreviewVendita('cv')"></div>
      <div class="row"><label>DICO/DIRI</label><input type="text" id="cv_dico" name="cv_dico" value="{{ cv.dico }}" oninput="updatePreviewVendita('cv')"></div>
      <div class="row"><label>Provvigione agenzia vendita %</label><input type="text" id="cv_provv_sale_pct" name="cv_provv_sale_pct" value="{{ cv.provv_sale_pct }}" oninput="updatePreviewVendita('cv')"></div>
      <div class="row"><label>Imprevisti</label><input type="text" id="cv_vendita_imprevisti" name="cv_vendita_imprevisti" value="{{ cv.vendita_imprevisti }}" oninput="updatePreviewVendita('cv')"></div>

      <div class="preview">
        <h4>Anteprima costi messa in vendita</h4>
        <div class="item"><span>Home staging</span><span id="cv_pv_hs_cost">—</span></div>
        <div class="item"><span>APE</span><span id="cv_pv_ape">—</span></div>
        <div class="item"><span>DICO/DIRI</span><span id="cv_pv_dico">—</span></div>
        <div class="item"><span>Provvigione agenzia</span><span id="cv_pv_provv">—</span></div>
        <div class="item"><span>Imprevisti</span><span id="cv_pv_vimp">—</span></div>
        <hr style="border:none;border-top:1px solid #1f2937;margin:6px 0"/>
        <div class="item"><b>Totale</b><b id="cv_pv_tot_vendita">—</b></div>
      </div>
      <div class="actions">
        <button class="btn primary" type="submit">Calcola</button>
        <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
      </div>
    </div>

    <!-- CV: NUOVO VALORE -->
    <div id="cv_val" class="tabcontent">
      <div class="row"><label>Street price (prezzo giusto di vendita)</label><input type="text" id="cv_street_price" name="cv_street_price" value="{{ cv.street_price }}" oninput="updatePreviewVal('cv');updatePreviewVendita('cv');"></div>
      <div class="row"><label>Incremento Home Staging %</label><input type="text" id="cv_inc_hs_pct" name="cv_inc_hs_pct" value="{{ cv.inc_hs_pct }}" oninput="updatePreviewVal('cv')"></div>
      <div class="row"><label>Incremento da ristrutturazione %</label><input type="text" id="cv_inc_ristr_pct" name="cv_inc_ristr_pct" value="{{ cv.inc_ristr_pct }}" oninput="updatePreviewVal('cv')"></div>

      <div class="preview">
        <h4>Anteprima valore</h4>
        <div class="item"><span>Valore finale percepito</span><span id="cv_pv_val_finale">—</span></div>
      </div>

      <div class="actions">
        <button class="btn primary" type="submit">Calcola</button>
        <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
      </div>
    </div>
  </form>
</div>

{% if results_cv %}
<div class="wrap card results">
  <h2>{{ results_cv.titolo or 'Riepilogo (Compravendita)' }}</h2>
  <div class="grid3">
    <div class="pill"><b>Tipo proprietà:</b> {{ results_cv.tipo_label }}</div>
    <div class="pill"><b>Valore catastale:</b> {{ results_cv.valore_catastale }}</div>
    <div class="pill"><b>Imposta di registro:</b> {{ results_cv.imposta_registro }}</div>
    <div class="pill"><b>Costo ristrutturazione:</b> {{ results_cv.ristrutturazione }}</div>
    <div class="pill"><b>Totale costi acquisto:</b> {{ results_cv.totale_acquisto }}</div>
    <div class="pill"><b>Costi messa in vendita:</b> {{ results_cv.costi_vendita }}</div>
    <div class="pill"><b>Valore finale percepito:</b> {{ results_cv.valore_finale }}</div>
    <div class="pill {{ results_cv.roi_class }}"><b>ROI:</b> {{ results_cv.roi }}</div>
  </div>

  <form method="post" action="/download_cv" style="margin-top:12px">
    {% for k,v in cv.items() %}
      <input type="hidden" name="{{k}}" value="{{v}}">
    {% endfor %}
    <button class="btn secondary" type="submit">Scarica Excel</button>
    <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
  </form>
</div>
{% endif %}

<!-- ========================== AFFITTI ========================== -->
<div class="wrap card parent-section" id="parent-affitti" style="display:none">
  <form method="post" id="form_affitti">
    <input type="hidden" name="parent_tab" value="affitti"/>
    <input type="hidden" id="active_tab_af" name="active_tab" value="{{ active_tab_af or 'af_acq' }}"/>

    <div class="tabs">
      <button type="button" class="tablink" data-tab="af_acq" onclick="openChild('affitti','af_acq')">Acquisto</button>
      <button type="button" class="tablink" data-tab="af_spese" onclick="openChild('affitti','af_spese')">Spese & Canone</button>
      <button type="button" class="tablink" data-tab="af_riep" onclick="openChild('affitti','af_riep')">Riepilogo Affitto</button>
    </div>

    <!-- AF: ACQUISTO (per locazione) -->
    <div id="af_acq" class="tabcontent">
      <div class="row"><label>Titolo operazione</label><input type="text" name="af_titolo" value="{{ af.titolo }}"></div>

      <div class="row"><label>Acquisto immobile</label><input type="text" name="af_ask" value="{{ af.ask }}"></div>
      <div class="row"><label>Imposta ipotecaria</label><input type="text" name="af_ipotecaria" value="{{ af.ipotecaria }}"></div>
      <div class="row"><label>Imposta catastale</label><input type="text" name="af_catastale" value="{{ af.catastale }}"></div>
      <div class="row"><label>Imposta di registro</label><input type="text" name="af_imp_reg" value="{{ af.imp_reg }}"></div>
      <div class="row"><label>Provvigioni agenzia (acquisto)</label><input type="text" name="af_agenzia" value="{{ af.agenzia }}"></div>
      <div class="row"><label>Ristrutturazione</label><input type="text" name="af_ristr" value="{{ af.ristr }}"></div>
      <div class="row"><label>Altre spese iniziali</label><input type="text" name="af_altre" value="{{ af.altre }}"></div>

      <div class="actions">
        <button class="btn primary" type="submit">Calcola</button>
        <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
      </div>
    </div>

    <!-- AF: SPESE & CANONE -->
    <div id="af_spese" class="tabcontent">
      <div class="row"><label>Canone mensile</label><input type="text" name="af_canone_mensile" value="{{ af.canone_mensile }}"></div>
      <div class="row"><label>Mesi locati/anno (es. 11 per 1 mese vacancy)</label><input type="text" name="af_mesi_locati" value="{{ af.mesi_locati }}"></div>
      <div class="row"><label>Spese condominiali a carico proprietario (mensili)</label><input type="text" name="af_spese_cond_mensili" value="{{ af.spese_cond_mensili }}"></div>
      <div class="row"><label>IMU annua</label><input type="text" name="af_imu_annua" value="{{ af.imu_annua }}"></div>
      <div class="row"><label>Assicurazione annua</label><input type="text" name="af_assicurazione_annua" value="{{ af.assicurazione_annua }}"></div>
      <div class="row"><label>Gestione/Amministrazione annua</label><input type="text" name="af_gestione_annua" value="{{ af.gestione_annua }}"></div>
      <div class="row"><label>Manutenzione % su ricavi lordi</label><input type="text" name="af_manut_pct" value="{{ af.manut_pct }}"></div>
      <div class="row"><label>Altre spese annue</label><input type="text" name="af_altre_annue" value="{{ af.altre_annue }}"></div>

      <div class="actions">
        <button class="btn primary" type="submit">Calcola</button>
        <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
      </div>
    </div>

    <!-- AF: RIEPILOGO AFFITTO -->
    <div id="af_riep" class="tabcontent">
      <div class="row">
        <label>Includi mutuo (da tab Mutuo)</label>
        <select name="af_includi_mutuo">
          <option value="no" {% if af.includi_mutuo=='no' %}selected{% endif %}>No</option>
          <option value="si" {% if af.includi_mutuo=='si' %}selected{% endif %}>Sì</option>
        </select>
      </div>

      <div class="preview">
        <h4>Indicatori</h4>
        <div class="item"><span>Ricavi lordi annui</span><span>{{ af_preview.ricavi_lordi }}</span></div>
        <div class="item"><span>Spese annue</span><span>{{ af_preview.spese_annue }}</span></div>
        <div class="item"><span>Reddito operativo (Ricavi - Spese)</span><span>{{ af_preview.noi }}</span></div>
        <div class="item"><span>Investimento totale</span><span>{{ af_preview.invest_tot }}</span></div>
        <div class="item"><span>ROI (senza debito)</span><span>{{ af_preview.roi }}</span></div>
        <hr style="border:none;border-top:1px solid #1f2937;margin:6px 0"/>
        <div class="item"><span>Mutuo - rata annua</span><span>{{ af_preview.mutuo_annuo }}</span></div>
        <div class="item"><span>Cashflow (Reddito operativo - rata)</span><span>{{ af_preview.cashflow }}</span></div>
        <div class="item"><span>Equity (inv. - capitale mutuato)</span><span>{{ af_preview.equity }}</span></div>
        <div class="item"><span>ROE (cashflow / equity)</span><span>{{ af_preview.roe }}</span></div>
        <hr style="border:none;border-top:1px solid #1f2937;margin:6px 0"/>
        <div class="item"><span>Payback mesi (senza debito, su Reddito operativo)</span><span>{{ af_preview.payback_mesi_no_debito }}</span></div>
        <div class="item"><span>Payback mesi (con debito, su Cashflow)</span><span>{{ af_preview.payback_mesi_con_debito }}</span></div>
      </div>

      {% if proiezione_5y and proiezione_5y|length > 0 %}
      <div class="preview">
        <h4>Proiezione 5 anni (cumulata)</h4>
        <table>
          <thead>
            <tr>
              <th>Anno</th>
              <th>Reddito operativo</th>
              <th>Cashflow</th>
              <th>ROI cumulato</th>
              <th>ROE cumulato</th>
            </tr>
          </thead>
          <tbody>
            {% for row in proiezione_5y %}
            <tr>
              <td>{{ row['Anno'] }}</td>
              <td>{{ row['Reddito operativo'] }}</td>
              <td>{{ row['Cashflow'] }}</td>
              <td>{{ row['ROI_cum'] }}</td>
              <td>{{ row['ROE_cum'] }}</td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      {% endif %}

      <div class="actions">
        <button class="btn primary" type="submit">Calcola</button>
        <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
      </div>
    </div>
  </form>
</div>

{% if results_af %}
<div class="wrap card results">
  <h2>{{ results_af.titolo or 'Riepilogo (Affitti)' }}</h2>
  <div class="grid3">
    <div class="pill"><b>Ricavi lordi annui:</b> {{ results_af.ricavi_lordi }}</div>
    <div class="pill"><b>Spese annue:</b> {{ results_af.spese_annue }}</div>
    <div class="pill"><b>Reddito operativo:</b> {{ results_af.noi }}</div>
    <div class="pill"><b>Investimento totale:</b> {{ results_af.invest_tot }}</div>
    <div class="pill"><b>ROI (senza debito):</b> {{ results_af.roi }}</div>
    <div class="pill"><b>Mutuo (rata annua):</b> {{ results_af.mutuo_annuo }}</div>
    <div class="pill"><b>Cashflow:</b> {{ results_af.cashflow }}</div>
    <div class="pill"><b>Equity:</b> {{ results_af.equity }}</div>
    <div class="pill"><b>ROE:</b> {{ results_af.roe }}</div>
    <div class="pill"><b>Payback (Reddito operativo):</b> {{ results_af.payback_no_debito }}</div>
    <div class="pill"><b>Payback (Cashflow):</b> {{ results_af.payback_con_debito }}</div>
  </div>

  <form method="post" action="/download_af" style="margin-top:12px">
    {% for k,v in af.items() %}
      <input type="hidden" name="{{k}}" value="{{v}}">
    {% endfor %}
    {% for k,v in mutuo.items() %}
      <input type="hidden" name="{{k}}" value="{{v}}">
    {% endfor %}
    <button class="btn secondary" type="submit">Scarica Excel</button>
    <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
  </form>
</div>
{% endif %}

<!-- ========================== MUTUO (GENITORE) ========================== -->
<div class="wrap card parent-section" id="parent-mutuo" style="display:none">
  <form method="post" id="form_mutuo">
    <input type="hidden" name="parent_tab" value="mutuo"/>

    <div class="row"><label>Capitale mutuato</label><input type="text" name="m_capitale" value="{{ mutuo.capitale }}"></div>
    <div class="row"><label>Tasso annuo %</label><input type="text" name="m_tasso" value="{{ mutuo.tasso }}"></div>
    <div class="row"><label>Durata (anni)</label><input type="text" name="m_anni" value="{{ mutuo.anni }}"></div>

    <div class="preview">
      <h4>Anteprima mutuo (anno 1)</h4>
      <div class="item"><span>Rata mensile</span><span>{{ mutuo_preview.rata_mensile }}</span></div>
      <div class="item"><span>Esborso annuo</span><span>{{ mutuo_preview.esborso_annuo }}</span></div>
      <div class="item"><span>Interessi anno 1</span><span>{{ mutuo_preview.interesse_anno1 }}</span></div>
      <div class="item"><span>Capitale rimborsato anno 1</span><span>{{ mutuo_preview.capitale_anno1 }}</span></div>
      <div class="item"><span>Residuo fine anno 1</span><span>{{ mutuo_preview.residuo_fine_anno1 }}</span></div>
    </div>

    <div class="actions">
      <button class="btn primary" type="submit">Calcola</button>
      <a class="btn secondary" href="{{ url_for('reset') }}">Reset</a>
    </div>
  </form>
</div>

<script>
/* ------- Parent tabs ------- */
function openParent(which){
  document.querySelectorAll(".parent-section").forEach(el=>el.style.display="none");
  document.querySelectorAll(".parent-btn").forEach(el=>el.classList.remove("active"));
  document.getElementById("parent-"+which).style.display="block");
  const btn=document.querySelector('.parent-btn[data-parent="'+which+'"]');
  if(btn) btn.classList.add("active");
}

/* ------- Child tabs ------- */
function openChild(parentKey, tabId){
  const parentEl = document.getElementById("parent-"+parentKey);
  parentEl.querySelectorAll(".tabcontent").forEach(el=>el.style.display="none");
  parentEl.querySelectorAll(".tablink").forEach(el=>el.classList.remove("active"));
  parentEl.querySelector('.tablink[data-tab="'+tabId+'"]').classList.add("active");
  parentEl.querySelector("#"+tabId).style.display="block";
  if(parentKey==='compravendita'){
    document.getElementById('active_tab_cv').value = tabId;
  } else if(parentKey==='affitti'){
    document.getElementById('active_tab_af').value = tabId;
  }
}

/* ------- Helpers ------- */
function num(v){ if(!v) return 0; v=(""+v).replace(",","."); return parseFloat(v)||0; }
function fmt(n){ return Math.round(n).toLocaleString('it-IT'); }

/* CV previews */
function autoFillByTipo(prefix){
  const tipo=document.getElementById(prefix+'_tipo').value;
  if(tipo==='prima'){document.getElementById(prefix+'_coeff').value="115.5";document.getElementById(prefix+'_imposta_pct').value="2";}
  else{document.getElementById(prefix+'_coeff').value="126";document.getElementById(prefix+'_imposta_pct').value="9";}
  updatePreviewCat(prefix);
}
function updatePreviewCat(prefix){
  const r=num(document.getElementById(prefix+'_rendita').value);
  const c=num(document.getElementById(prefix+'_coeff').value);
  const p=num(document.getElementById(prefix+'_imposta_pct').value);
  const val=r*c; const imp=val*(p/100);
  document.getElementById(prefix+'_pv_val_cat').innerHTML=fmt(val);
  document.getElementById(prefix+'_pv_imp_reg').innerHTML=fmt(imp);
}
function updatePreviewVendita(prefix){
  const spEl=document.getElementById(prefix+"_street_price");
  const sp = spEl ? num(spEl.value) : 0;
  const hsPercent=num(document.querySelector("[name='"+prefix+"_hs_percent']").value);
  const ape=num(document.getElementById(prefix+"_ape").value);
  const dico=num(document.getElementById(prefix+"_dico").value);
  const provvPct=num(document.getElementById(prefix+"_provv_sale_pct").value);
  const vimp=num(document.getElementById(prefix+"_vendita_imprevisti").value);
  const hsCost=sp*(hsPercent/100);
  const provv=sp*(provvPct/100);
  const totale=hsCost+ape+dico+provv+vimp;
  document.getElementById(prefix+"_pv_hs_cost").innerHTML=fmt(hsCost);
  document.getElementById(prefix+"_pv_ape").innerHTML=fmt(ape);
  document.getElementById(prefix+"_pv_dico").innerHTML=fmt(dico);
  document.getElementById(prefix+"_pv_provv").innerHTML=fmt(provv);
  document.getElementById(prefix+"_pv_vimp").innerHTML=fmt(vimp);
  document.getElementById(prefix+"_pv_tot_vendita").innerHTML=fmt(totale);
}
function updatePreviewVal(prefix){
  const sp=num(document.getElementById(prefix+'_street_price').value);
  const incHS=num(document.getElementById(prefix+'_inc_hs_pct').value);
  const incR=num(document.getElementById(prefix+'_inc_ristr_pct').value);
  const valFinale = sp * (1 + incHS/100 + incR/100);
  document.getElementById(prefix+'_pv_val_finale').innerHTML=fmt(valFinale);
}

/* Init */
document.addEventListener('DOMContentLoaded',()=>{
  // default parent
  openParent("{{ parent_tab or 'compravendita' }}");

  // default child tabs
  openChild('compravendita', "{{ active_tab_cv or 'cv_acq' }}");
  openChild('affitti', "{{ active_tab_af or 'af_acq' }}");

  // initial previews
  updatePreviewCat('cv');
  updatePreviewVendita('cv');
  updatePreviewVal('cv');
});
</script>
</body>
</html>
"""

# -------------------- calcoli backend --------------------
def compute_cv(form):
    titolo = (form.get("cv_titolo") or "").strip()

    # Acquisto
    ask = parse_num(form.get("cv_ask"))
    ipotecaria = parse_num(form.get("cv_ipotecaria"))
    catastale_cost = parse_num(form.get("cv_catastale"))
    agenzia = parse_num(form.get("cv_agenzia"))
    architetto = parse_num(form.get("cv_architetto"))
    condono = parse_num(form.get("cv_condono"))
    condominio = parse_num(form.get("cv_condominio"))
    utenze = parse_num(form.get("cv_utenze"))
    imprevisti = parse_num(form.get("cv_imprevisti"))
    ristrut_tipo = form.get("cv_ristrut_tipo","nessuna")
    perc_map = {"nessuna":0,"piccola":0.10,"intermedia":0.20,"complessa":0.60}
    ristr_perc = perc_map.get(ristrut_tipo,0.0)
    ristrutturazione = ask * ristr_perc

    # Catastale & Registro
    tipo = form.get("cv_tipo","prima")
    rendita = parse_num(form.get("cv_rendita"))
    coeff = parse_num(form.get("cv_coeff"))
    imposta_pct = parse_num(form.get("cv_imposta_pct"))
    valore_catastale = rendita * coeff
    imposta_registro = valore_catastale * (imposta_pct / 100.0)

    totale_acquisto = (ask + ipotecaria + catastale_cost + agenzia + architetto +
                       condono + condominio + utenze + imprevisti + ristrutturazione +
                       imposta_registro)

    # Vendita (costi messa in vendita)
    street_price = parse_num(form.get("cv_street_price"))
    hs_percent = parse_num(form.get("cv_hs_percent"))
    ape = parse_num(form.get("cv_ape"))
    dico = parse_num(form.get("cv_dico"))
    provv_sale_pct = parse_num(form.get("cv_provv_sale_pct"))
    vendita_imprevisti = parse_num(form.get("cv_vendita_imprevisti"))

    hs_cost = street_price * (hs_percent/100.0)
    provv_sale_cost = street_price * (provv_sale_pct/100.0)
    costi_vendita = hs_cost + ape + dico + provv_sale_cost + vendita_imprevisti

    # Nuovo Valore
    inc_hs_pct = parse_num(form.get("cv_inc_hs_pct"))
    inc_ristr_pct = parse_num(form.get("cv_inc_ristr_pct"))
    valore_finale = street_price * (1 + inc_hs_pct/100.0 + inc_ristr_pct/100.0)

    # ROI
    if totale_acquisto > 0:
        roi_val = (valore_finale - totale_acquisto) / totale_acquisto * 100.0
    else:
        roi_val = 0.0
    roi_class = "roi-good" if roi_val > 30 else ""

    results = {
        "titolo": titolo if titolo else None,
        "tipo_label": "Prima casa" if tipo=="prima" else "Seconda casa",
        "valore_catastale": f"{round(valore_catastale):,}".replace(",", "."),
        "imposta_registro": f"{round(imposta_registro):,}".replace(",", "."),
        "ristrutturazione": f"{round(ristrutturazione):,}".replace(",", "."),
        "totale_acquisto": f"{round(totale_acquisto):,}".replace(",", "."),
        "costi_vendita": f"{round(costi_vendita):,}".replace(",", "."),
        "valore_finale": f"{round(valore_finale):,}".replace(",", "."),
        "roi": f"{roi_val:.1f}%",
        "roi_class": roi_class,
    }

    # formvals (per Excel)
    cv = {k: (form.get(k) or "") for k in form.keys() if k.startswith("cv_")}
    cv.update({
        "cv_imposta_registro": f"{round(imposta_registro):,}".replace(",", "."),
        "cv_ristrutturazione": f"{round(ristrutturazione):,}".replace(",", "."),
    })
    return results, cv

def compute_mutuo(form):
    capitale = parse_num(form.get("m_capitale"))
    tasso = parse_num(form.get("m_tasso"))
    anni = parse_num(form.get("m_anni"))
    rata_m, esborso_annuo, interesse_1, capitale_1, residuo = amortize_first_year(capitale, tasso, anni)
    m = {
        "capitale": str(int(round(capitale))) if capitale else "0",
        "tasso": str(tasso),
        "anni": str(int(anni)) if anni else "0",
        "rata_mensile": f"{round(rata_m):,}".replace(",", "."),
        "esborso_annuo": f"{round(esborso_annuo):,}".replace(",", "."),
        "interesse_anno1": f"{round(interesse_1):,}".replace(",", "."),
        "capitale_anno1": f"{round(capitale_1):,}".replace(",", "."),
        "residuo_fine_anno1": f"{round(residuo):,}".replace(",", "."),
    }
    return m

def compute_affitti(form, mutuo_ctx):
    titolo = (form.get("af_titolo") or "").strip()

    # Investimento iniziale (affitto)
    ask = parse_num(form.get("af_ask"))
    ipotecaria = parse_num(form.get("af_ipotecaria"))
    catastale = parse_num(form.get("af_catastale"))
    imp_reg = parse_num(form.get("af_imp_reg"))
    agenzia = parse_num(form.get("af_agenzia"))
    ristr = parse_num(form.get("af_ristr"))
    altre = parse_num(form.get("af_altre"))
    invest_tot = ask + ipotecaria + catastale + imp_reg + agenzia + ristr + altre

    # Canone & Spese
    canone_mensile = parse_num(form.get("af_canone_mensile"))
    mesi_locati = parse_num(form.get("af_mesi_locati"), 12.0)
    spese_cond_mensili = parse_num(form.get("af_spese_cond_mensili"))
    imu_annua = parse_num(form.get("af_imu_annua"))
    assicurazione = parse_num(form.get("af_assicurazione_annua"))
    gestione = parse_num(form.get("af_gestione_annua"))
    manut_pct = parse_num(form.get("af_manut_pct"))
    altre_annue = parse_num(form.get("af_altre_annue"))

    ricavi_lordi = canone_mensile * mesi_locati
    spese_cond_annue = spese_cond_mensili * 12.0
    manut_annua = ricavi_lordi * (manut_pct / 100.0)
    spese_annue = spese_cond_annue + imu_annua + assicurazione + gestione + manut_annua + altre_annue
    noi = ricavi_lordi - spese_annue

    # ROI senza debito
    roi = (noi / invest_tot * 100.0) if invest_tot > 0 else 0.0

    # Mutuo (opzionale)
    includi_mutuo = (form.get("af_includi_mutuo") or "no").lower() == "si"
    # mutuo_ctx valori sono stringhe formattate -> puliamo
    mutuo_annuo = 0.0
    capitale_mutuato = 0.0
    try:
        mutuo_annuo = float(mutuo_ctx.get("esborso_annuo","0").replace(".","").replace(",","."))
    except:
        mutuo_annuo = 0.0
    try:
        capitale_mutuato = float(mutuo_ctx.get("capitale","0").replace(".","").replace(",","."))
    except:
        capitale_mutuato = 0.0

    cashflow = noi - (mutuo_annuo if includi_mutuo else 0.0)
    equity = max(invest_tot - (capitale_mutuato if includi_mutuo else 0.0), 0.0)
    roe = (cashflow / equity * 100.0) if equity > 0 else 0.0

    # Payback (mesi)
    noi_mensile = noi / 12.0
    cashflow_mensile = cashflow / 12.0
    payback_no_debito = "N/D"
    payback_con_debito = "N/D"
    if invest_tot > 0 and noi_mensile > 0:
        mesi = math.ceil(invest_tot / noi_mensile)
        payback_no_debito = f"{mesi} mesi (~{mesi/12:.1f} anni)"
    if equity > 0 and cashflow_mensile > 0 and includi_mutuo:
        mesi = math.ceil(equity / cashflow_mensile)
        payback_con_debito = f"{mesi} mesi (~{mesi/12:.1f} anni)"

    # Proiezione 5 anni (costanti, senza crescita)
    proiezione = []
    cum_noi = 0.0
    cum_cf = 0.0
    for year in range(1, 6):
        cum_noi += noi
        cum_cf += cashflow
        roi_cum = (cum_noi / invest_tot * 100.0) if invest_tot > 0 else 0.0
        roe_cum = (cum_cf / equity * 100.0) if equity > 0 else 0.0
        proiezione.append({
            "Anno": year,
            "Reddito operativo": f"{round(noi):,}".replace(",", "."),
            "Cashflow": f"{round(cashflow):,}".replace(",", "."),
            "ROI_cum": f"{roi_cum:.1f}%",
            "ROE_cum": f"{roe_cum:.1f}%",
        })

    results = {
        "titolo": titolo if titolo else None,
        "ricavi_lordi": f"{round(ricavi_lordi):,}".replace(",", "."),
        "spese_annue": f"{round(spese_annue):,}".replace(",", "."),
        "noi": f"{round(noi):,}".replace(",", "."),
        "invest_tot": f"{round(invest_tot):,}".replace(",", "."),
        "roi": f"{roi:.1f}%",
        "mutuo_annuo": f"{round(mutuo_annuo):,}".replace(",", ".") if includi_mutuo else "0",
        "cashflow": f"{round(cashflow):,}".replace(",", "."),
        "equity": f"{round(equity):,}".replace(",", "."),
        "roe": f"{roe:.1f}%",
        "payback_no_debito": payback_no_debito,
        "payback_con_debito": payback_con_debito,
    }

    # Anteprima per pannello
    preview = {
        "ricavi_lordi": results["ricavi_lordi"],
        "spese_annue": results["spese_annue"],
        "noi": results["noi"],
        "invest_tot": results["invest_tot"],
        "roi": results["roi"],
        "mutuo_annuo": results["mutuo_annuo"],
        "cashflow": results["cashflow"],
        "equity": results["equity"],
        "roe": results["roe"],
        "payback_mesi_no_debito": results["payback_no_debito"],
        "payback_mesi_con_debito": results["payback_con_debito"],
    }

    # stato per Excel
    af = {k: (form.get(k) or "") for k in form.keys() if k.startswith("af_")}
    return results, af, preview, proiezione

# -------------------- routes --------------------
@app.route("/", methods=["GET","POST"])
def index():
    # default cv
    cv = {
        "titolo":"", "ask":"150000","ipotecaria":"50","catastale":"50","agenzia":"3000","architetto":"2000",
        "condono":"0","condominio":"0","utenze":"500","imprevisti":"2000",
        "ristrut_tipo":"nessuna","ristrutturazione":"0",
        "tipo":"prima","rendita":"500","coeff":"115.5","imposta_pct":"2","imposta_registro":"0",
        "hs_percent":"2","ape":"200","dico":"250","provv_sale_pct":"3","vendita_imprevisti":"500",
        "street_price":"220000","inc_hs_pct":"5","inc_ristr_pct":"10"
    }
    # default affitti
    af = {
        "titolo":"",
        "ask":"120000","ipotecaria":"50","catastale":"50","imp_reg":"2000","agenzia":"2000","ristr":"10000","altre":"0",
        "canone_mensile":"850","mesi_locati":"11","spese_cond_mensili":"30","imu_annua":"600","assicurazione_annua":"150",
        "gestione_annua":"200","manut_pct":"5","altre_annue":"0","includi_mutuo":"no"
    }
    # default mutuo
    mutuo = {"capitale":"80000","tasso":"4.2","anni":"25",
             "rata_mensile":"0","esborso_annuo":"0","interesse_anno1":"0","capitale_anno1":"0","residuo_fine_anno1":"0"}

    parent_tab = request.form.get("parent_tab") if request.method=="POST" else "compravendita"
    active_tab_cv = "cv_acq"
    active_tab_af = "af_acq"

    results_cv = None
    results_af = None
    af_preview = {"ricavi_lordi":"—","spese_annue":"—","noi":"—","invest_tot":"—","roi":"—",
                  "mutuo_annuo":"—","cashflow":"—","equity":"—","roe":"—",
                  "payback_mesi_no_debito":"—","payback_mesi_con_debito":"—"}
    proiezione_5y = []

    # override da POST
    if request.method == "POST":
        # popolamento cv da form
        for k in list(cv.keys()):
            form_key = "cv_"+k
            if form_key in request.form:
                cv[k] = request.form.get(form_key)

        # popolamento affitti
        for k in list(af.keys()):
            form_key = "af_"+k
            if form_key in request.form:
                af[k] = request.form.get(form_key)

        # popolamento mutuo
        for k_form,k_state in [("m_capitale","capitale"),("m_tasso","tasso"),("m_anni","anni")]:
            if k_form in request.form:
                mutuo[k_state] = request.form.get(k_form)

        # smista
        if parent_tab == "compravendita":
            active_tab_cv = request.form.get("active_tab", "cv_acq")
            # prepara dizionario prefissato per funzione
            f = { "cv_"+k: v for k,v in cv.items() }
            results_cv, cv_state = compute_cv(f)
            cv["imposta_registro"] = cv_state["cv_imposta_registro"]
            cv["ristrutturazione"] = cv_state["cv_ristrutturazione"]

        elif parent_tab == "affitti":
            active_tab_af = request.form.get("active_tab", "af_acq")
            mutuo_ctx = compute_mutuo({
                "m_capitale": mutuo["capitale"],
                "m_tasso": mutuo["tasso"],
                "m_anni": mutuo["anni"],
            })
            f = { "af_"+k: v for k,v in af.items() }
            results_af, af_state, af_preview, proiezione_5y = compute_affitti(f, mutuo_ctx)
            mutuo.update(mutuo_ctx)

        elif parent_tab == "mutuo":
            mutuo = compute_mutuo(request.form)

    mutuo_preview = {
        "rata_mensile": mutuo.get("rata_mensile","0"),
        "esborso_annuo": mutuo.get("esborso_annuo","0"),
        "interesse_anno1": mutuo.get("interesse_anno1","0"),
        "capitale_anno1": mutuo.get("capitale_anno1","0"),
        "residuo_fine_anno1": mutuo.get("residuo_fine_anno1","0"),
    }

    return render_template_string(
        HTML,
        parent_tab=parent_tab,
        active_tab_cv=active_tab_cv,
        active_tab_af=active_tab_af,
        cv=cv, af=af, mutuo=mutuo,
        results_cv=results_cv, results_af=results_af,
        af_preview=af_preview, proiezione_5y=proiezione_5y,
        mutuo_preview=mutuo_preview
    )

# -------------------- download --------------------
@app.route("/download_cv", methods=["POST"])
def download_cv():
    results, cv_state = compute_cv(request.form)
    titolo = results.get("titolo") or "Operazione Compravendita"
    file_slug = slugify(titolo)

    df_inputs = pd.DataFrame([(k,v) for k,v in request.form.items() if k.startswith("cv_")], columns=["Voce","Valore"])
    df_results = pd.DataFrame([
        ["Titolo operazione", titolo],
        ["Tipo proprietà", results["tipo_label"]],
        ["Valore catastale", results["valore_catastale"]],
        ["Imposta di registro", results["imposta_registro"]],
        ["Costo ristrutturazione", results["ristrutturazione"]],
        ["Totale costi acquisto", results["totale_acquisto"]],
        ["Costi messa in vendita", results["costi_vendita"]],
        ["Valore finale percepito", results["valore_finale"]],
        ["ROI", results["roi"]],
    ], columns=["Risultato","Valore"])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_inputs.to_excel(writer, index=False, sheet_name="Input")
        df_results.to_excel(writer, index=False, sheet_name="Risultati")
        for _, ws in writer.sheets.items():
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    val = cell.value
                    l = len(str(val)) if val is not None else 0
                    if l > max_len: max_len = l
                ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len+2, 60))
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"{file_slug}_compravendita.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/download_af", methods=["POST"])
def download_af():
    # ricostruisci un contesto mutuo minimale per equity/cashflow nel ricalcolo
    mutuo_ctx = {
        "capitale": request.form.get("capitale","0"),
        "esborso_annuo": request.form.get("esborso_annuo","0"),
    }
    results, af_state, preview, proiezione = compute_affitti(request.form, mutuo_ctx)
    titolo = results.get("titolo") or "Operazione Affitti"
    file_slug = slugify(titolo)

    df_inputs = pd.DataFrame([(k,v) for k,v in request.form.items() if k.startswith("af_")], columns=["Voce","Valore"])
    df_mutuo  = pd.DataFrame([(k,v) for k,v in request.form.items() if k in ["capitale","esborso_annuo"]], columns=["Voce","Valore"])
    df_results = pd.DataFrame([
        ["Titolo operazione", titolo],
        ["Ricavi lordi annui", results["ricavi_lordi"]],
        ["Spese annue", results["spese_annue"]],
        ["Reddito operativo", results["noi"]],
        ["Investimento totale", results["invest_tot"]],
        ["ROI (senza debito)", results["roi"]],
        ["Mutuo (rata annua)", results["mutuo_annuo"]],
        ["Cashflow", results["cashflow"]],
        ["Equity", results["equity"]],
        ["ROE", results["roe"]],
        ["Payback Reddito operativo (mesi)", results["payback_no_debito"]],
        ["Payback Cashflow (mesi)", results["payback_con_debito"]],
    ], columns=["Risultato","Valore"])

    # Proiezione 5 anni in un foglio dedicato
    df_proj = pd.DataFrame(proiezione)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_inputs.to_excel(writer, index=False, sheet_name="Input_Affitti")
        df_mutuo.to_excel(writer, index=False, sheet_name="Mutuo")
        df_results.to_excel(writer, index=False, sheet_name="Risultati")
        if not df_proj.empty:
            df_proj.to_excel(writer, index=False, sheet_name="Proiezione_5_anni")

        # auto-fit
        for _, ws in writer.sheets.items():
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    val = cell.value
                    l = len(str(val)) if val is not None else 0
                    if l > max_len: max_len = l
                ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len+2, 60))
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"{file_slug}_affitti.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/reset")
def reset():
    return redirect(url_for("index"))

# -------------------- main --------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
